import json
import re
from pathlib import Path

import openpyxl
from openai import OpenAI

STUDENTS_FILE = "students.json"
EXCEL_FILE = "动图.xlsx"

OUTPUT_FILE = "动图_已填写.xlsx"

client = OpenAI(
    api_key=API_KEY,
    base_url=BASE_URL
)

def load_questions(workbook):
    """
    收集所有sheet中的题目
    默认：
        C列 = 题目内容
    """

    questions = []

    for ws in workbook.worksheets:

        for row in range(2, ws.max_row + 1):

            q = ws.cell(row=row, column=3).value

            if q is None:
                continue

            q = str(q).strip()

            if not q:
                continue

            questions.append(
                {
                    "sheet": ws.title,
                    "row": row,
                    "question": q
                }
            )

    return questions

def build_questionnaire(questions):

    txt = []

    for i, item in enumerate(questions, start=1):
        txt.append(f"{i}. {item['question']}")

    return "\n\n".join(txt)

def ask_qwen(profile, questionnaire_text, total_questions):

    system_prompt = f"""
你是一个真实的大学生，以下是你的个人信息：

{json.dumps(profile, ensure_ascii=False, indent=2)}

请充分代入角色。

把自己当成一个真实存在的人。

你拥有自己的成长经历、消费观、手机使用习惯、社交圈、学习状态和价值观。

允许根据画像合理补充细节背景，
但必须与画像保持一致。

针对问卷问题给出真实、自然、有个人特色的回答。

请保证：

- 回答符合年龄
- 回答符合专业
- 回答符合经济条件
- 回答符合MBTI
- 回答符合兴趣爱好
- 回答符合恋爱状态
- 回答符合设备使用习惯
- 回答前后一致

【强制规则】
1. 问卷共有 {total_questions} 道题。
   请返回长度必须等于 {total_questions} 的JSON。
2. 返回合法的JSON对象：
{{
  "answers": [
    "答案1",
    "答案2",
    ...
  ]
}}
不要输任何其他内容。
不要输出```json。
3. 不要重复问题
4. 选择题直接写：
   A、xxxx；C、xxxx；F、xxxx
5. 开放题请写真实自然的大学生表达
"""

    user_prompt = f"""
请回答以下问卷：

{questionnaire_text}
"""

    response = client.chat.completions.create(
        model=MODEL_NAME,
        temperature=1.0,
        top_p=0.95,
        messages=[
            {
                "role": "system",
                "content": system_prompt
            },
            {
                "role": "user",
                "content": user_prompt
            }
        ]
    )

    return response.choices[0].message.content.strip()

def parse_answers(answer_text):

    answer_text = answer_text.strip()

    answer_text = answer_text.replace("```json", "")
    answer_text = answer_text.replace("```", "")
    answer_text = answer_text.strip()

    data = json.loads(answer_text)

    return data["answers"]

def main():

    print("读取学生画像...")

    with open(STUDENTS_FILE, "r", encoding="utf-8") as f:
        students = json.load(f)

    print("读取问卷...")

    wb = openpyxl.load_workbook(EXCEL_FILE)

    questions = load_questions(wb)

    questionnaire_text = build_questionnaire(questions)

    total_questions = len(questions)

    print(f"题目数: {total_questions}")
    print(f"学生数: {len(students)}")

    # D列开始
    start_col = 4

    for idx, student in enumerate(students):

        col = start_col + idx

        student_name = student["name"]

        print("=" * 80)
        print(f"开始处理学生 {student_name}")

        try:

            answer_text = ask_qwen(
                profile=student,
                questionnaire_text=questionnaire_text,
                total_questions=total_questions
            )

            answers = parse_answers(answer_text)

            print(
                f"返回答案数: {len(answers)} / {total_questions}"
            )

            if len(answers) != total_questions:
                raise ValueError(
                    f"答案数量异常: {len(answers)} != {total_questions}"
                )

            # 写表头
            for ws in wb.worksheets:
                ws.cell(row=1, column=col).value = student_name

            # 写回对应sheet
            answer_index = 0

            for ws in wb.worksheets:

                for row in range(2, ws.max_row + 1):

                    q = ws.cell(row=row, column=3).value

                    if q is None:
                        continue

                    q = str(q).strip()

                    if not q:
                        continue

                    ws.cell(
                        row=row,
                        column=col
                    ).value = answers[answer_index]

                    answer_index += 1

            print(f"完成 {student_name}")

        except Exception as e:

            print(f"失败 {student_name}")
            print(e)

    wb.save(OUTPUT_FILE)

    print("=" * 80)
    print("全部完成")
    print("输出文件：", OUTPUT_FILE)


if __name__ == "__main__":
    main()
